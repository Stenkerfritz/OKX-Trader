# 1. Mainmethode
# 2. Funktionen für Eingabe der Grundwerte
# 3. Funktionen für die Berechnung des Depotwertes
# 4. Funktionen für den websocket-Zugriff
# 5. Funktionen für den Zugriff auf die API
# 6. Schleife um alles, was immer wieder ausgeführt werden soll


from openpyxl import Workbook, load_workbook
import config
import requests
from bs4 import BeautifulSoup
import okx.Account as Account
import okx.MarketData as MarketData
import okx.PublicData as PublicData
import okx.SpreadTrading as SpreadTrading
import funktion
import json
from datetime import datetime,timezone
import time
import okx.Trade as Trade


if __name__ == "__main__":
    #    print("Test")

    flag = "1"    # live trading: 0, demo trading: 1
    # API initialization
    apikey = config.GET_API_KEY(flag)
    secretkey = config.GET_SECRET_KEY(flag)
    passphrase = config.GET_PASS(flag)

    handelsBudget = input("Bitte geben Sie den Startbetrag ein!")
    abfrage = input("Moechten Sie ihren Gewinn in BTC oder USDT")
    steigung = input("Bitte geben Sie die Steigerungquote ein!")
    # Platz für Abfrage Verkauf in "BTC oder USDT"

    marketDataAPI = MarketData.MarketAPI(flag=flag)
    publicDataAPI = PublicData.PublicAPI(flag=flag)

    # stufen = funktion.sicherheit
    # quote = funktion.steigung

    # Get maximum buy/sell amount or open amount

    # spreadAPI = SpreadTrading.SpreadTradingAPI(apikey, secretkey, passphrase, False, flag)
    #marketDataAPI = MarketData.MarketAPI(flag=flag)
    # get tickers
    result = marketDataAPI.get_tickers(instType="SPOT")
    print(result["code"])
    print(result["msg"])
    print(result["data"])
    json.dump(result,open('results/api-resultat.json','w'),indent=4,sort_keys=True)

    zeit = time.time()
    print(int(zeit))

    #accountAPI = Account.AccountAPI(apikey, secretkey, passphrase, False, flag)

    # Get account balance
    #result = accountAPI.get_account_balance()
    #print(result)

    wb = Workbook()
    ws = wb.active
                                                    # hier kommt noch eine Schleife
    ws.titel = "OKX-Trade aktuell"
    ws["A1"].value = "Vorgangs NR"
    ws["B1"].value = "Waehrung"
    ws["C1"].value = "Kauf- und Verkaufs-Kurs"
    ws["D1"].value = "Betrag Kauf/Verkauf"
    ws["E1"].value = "gekaufte Menge"
    ws["F1"].value = "Kosten"
    ws["G1"].value = "tatsaechliche Menge"
    ws["H1"].value = "Druchschnittskurs"
    ws["I1"].value = "Gesamt Menge"
    ws["J1"].value = "Gesamt Kapital"
    ws["K1"].value = "Aktion"
    wb.save("OKX-Trade.xlsx")

    menge = 0
    j = 0
    i = 1

    kauf = funktion.GET_KAUF(flag, handelsBudget)

    result = marketDataAPI.get_ticker(instId="BTC-USDT")
    json.dump(result, open('results/api-resultat-BTC-USDT.json', 'w'), indent=4, sort_keys=True)
    handelsWaehrung = result["data"][0]["instId"]
    handelsKurs = result["data"][0]["last"]
    print(handelsWaehrung)
    print(handelsKurs)
    budget = str(float(handelsBudget) / float(handelsKurs))

    ws["A2"].value = str(i) + "." + str(j + 1)
    ws["B2"].value = handelsWaehrung
    ws["C2"].value = handelsKurs
    ws["D2"].value = str(handelsBudget)
    gesamtMenge = str(float(handelsBudget) * 1 / float(handelsKurs))
    ws["E2"].value = gesamtMenge
    kosten = float(gesamtMenge) * 0.108 / 100
    ws["F2"].value = str(kosten)
    tatsaechlicheMenge = float(gesamtMenge) - float(kosten)
    ws["G2"].value = str(tatsaechlicheMenge)
    druchschnittskurs = float(handelsBudget)/tatsaechlicheMenge
    ws["H2"].value = str(druchschnittskurs)
    ws["I2"].value = str(ws["G2"].value)
    ws["J2"].value = str(ws["D2"].value)
    ws["K2"].value = "Kauf"

    wb.save("OKX-Trade.xlsx")
    j = j + 1

    while j != 0:
        try:
            time.sleep(121.5)
            marketDataAPI = MarketData.MarketAPI(flag=flag)
            result = marketDataAPI.get_ticker(instId="BTC-USDT")
            handelsKurs = result["data"][0]["last"]
            print(handelsKurs)

            if j != 3: #druchschnittskurs * 0.99 > float(handelsKurs):
                kauf = funktion.GET_KAUF(flag, handelsBudget)
                ws.insert_rows(2)
                ws["A2"].value = str(i) + "." + str(j + 1)
                ws["B2"].value = handelsWaehrung
                ws["C2"].value = handelsKurs
                ws["D2"].value = str(handelsBudget)
                gesamtMenge = float(handelsBudget) * 1 / float(handelsKurs)
                ws["E2"].value = gesamtMenge
                kosten = float(gesamtMenge) * 0.108 / 100
                ws["F2"].value = str(kosten)
                tatsaechlicheMenge = float(gesamtMenge) - kosten
                ws["G2"].value = str(tatsaechlicheMenge)
                gesamtBudget = handelsBudget + float(ws["J3"].value)
                eingekauteMenge = tatsaechlicheMenge + float(ws["I3"].value)
                ws["I2"].value = str(eingekauteMenge)
                ws["J2"].value = str(gesamtBudget)
                druchschnittskurs = float(ws["J2"].value) / float(ws["I2"].value)
                ws["H2"].value = str(druchschnittskurs)
                ws["K2"].value = "Kauf"
                wb.save("OKX-Trade.xlsx")
                j = j + 1
            elif (float(ws["H2"].value) * 1.015 < float(handelsKurs)) and (abfrage.casefold() == "BTC".casefold()):
                verkauf = funktion.verkaufBtc(flag, str(ws["I2"].value))
                ws.insert_rows(2)
                ws["A2"].value = str(i) + "." + str(j + 1)
                ws["B2"].value = handelsWaehrung
                ws["C2"].value = handelsKurs
                ws["D2"].value = ws["J3"].value
                # gesamtMenge = float(handelsBudget) * 1 / float(handelsKurs)
                # ws["E2"].value = gesamtMenge
                # kosten = gesamtMenge * 0.108 / 100
                # ws["F2"].value = kosten
                #tatsaechlicheMenge = gesamtMenge - kosten
                #ws["G2"].value = tatsaechlicheMenge
                gesamtBudget = float(handelsBudget) + float(str(ws["J3"].value))
                eingekauteMenge = tatsaechlicheMenge + float(str(ws["I3"].value))
                ws["I2"].value = eingekauteMenge
                ws["J2"].value = ws["J3"].value
                ws["H2"].value = str(float(ws["H3"].value)*float(ws["J3"].value))
                ws["K2"].value = "verkauf"
                ws.insert_rows(1)
                ws.insert_rows(1)
                ws.insert_rows(2)
                j = 0
                i = i + 1
                wb.save("OKX-Trade.xlsx")
            elif (float(ws["H2"].value) * 1.015 < float(handelsKurs)) and (abfrage.casefold() == "USDT".casefold()):
                verkauf = funktion.verkaufUsdt(flag, str(ws["I2"].value))
                ws.insert_rows(2)
                ws["A2"].value = str(i) + "." + str(j + 1)
                ws["B2"].value = handelsWaehrung
                ws["C2"].value = handelsKurs
                ws["D2"].value = ws["J3"].value
                # gesamtMenge = float(handelsBudget) * 1 / float(handelsKurs)
                # ws["E2"].value = gesamtMenge
                # kosten = gesamtMenge * 0.108 / 100
                # ws["F2"].value = kosten
                #tatsaechlicheMenge = gesamtMenge - kosten
                #ws["G2"].value = tatsaechlicheMenge
                gesamtBudget = float(handelsBudget) + float(str(ws["J3"].value))
                eingekauteMenge = tatsaechlicheMenge + float(str(ws["I3"].value))
                ws["I2"].value = eingekauteMenge
                ws["J2"].value = ws["J3"].value
                ws["H2"].value = str(float(ws["H3"].value)*float(ws["J3"].value))
                ws["K2"].value = "verkauf"
                ws.insert_rows(1)
                ws.insert_rows(1)
                ws.insert_rows(2)
                j = 0
                i = i + 1
                wb.save("OKX-Trade.xlsx")
            else:
                print("NIX")

                #if für Gewinn in BTC erzielen
                # else für Gewinn in USDT erzielen

        except Exception as e:
            print(f"Fehler aufgetreten: {e}")
            continue

